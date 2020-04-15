from openpyxl import load_workbook


class HandleExcel:
    def __init__(self, path=None):
        if path is None:
            self.path = "../case/base.xlsx"
        self.instance = load_workbook(self.path)

    def get_target_sheet(self, index):
        return self.instance[self.instance.sheetnames[index]]

    # 行号row_index从1开始
    def get_row_value_from_target_sheet(self, row_index=1, target_index=0):
        row_values = []
        sheet = self.get_target_sheet(target_index)
        for v in sheet[row_index]:
            row_values.append(v.value)
        return row_values

    def get_row_nums(self,target_index=0):
        sheet = self.get_target_sheet(target_index)
        return sheet.max_row


if __name__ == "__main__":
    handleExcel = HandleExcel()

    # 行是从 1 开始的，但是又要去掉表头，所以从 2 开始
    for i in range(2, handleExcel.get_row_nums() + 1):
        # 获取每一行的值
        print(handleExcel.get_row_value_from_target_sheet(i))
