from openpyxl import load_workbook
import os


class Excel:

    def __init__(self, path=None):
        if path is None:
            self.path = os.path.dirname(__file__) + "/../case/base.xlsx"
        self.instance = load_workbook(self.path)

    def get_target_sheet(self, index):
        return self.instance[self.instance.sheetnames[index]]

    # 行号row_index从1开始
    def get_row_value_from_target_sheet(self, row_index=1, target_index=0):
        row_values = []
        sheet = self.get_target_sheet(target_index)
        for v in sheet[row_index]:
            row_values.append(v.value)
        return row_values

    # 回写了值之后需要执行 save 方法
    def set_cell_value_from_target_sheet(self, row_index, cell_index, value, target_index=0):
        # self.instance.active
        sheet = self.get_target_sheet(target_index)
        sheet.cell(row_index, cell_index, value)

    def save(self, path=None):
        if path is None:
            path = self.path
        self.instance.save(path)

    def get_row_nums(self, target_index=0):
        sheet = self.get_target_sheet(target_index)
        return sheet.max_row


if __name__ == "__main__":
    handleExcel = Excel()
    handleExcel.set_cell_value_from_target_sheet(2, 3, "测试啊测试")
    # 行是从 1 开始的，但是又要去掉表头，所以从 2 开始
    for i in range(2, handleExcel.get_row_nums() + 1):
        # 获取每一行的值
        print(handleExcel.get_row_value_from_target_sheet(i))
